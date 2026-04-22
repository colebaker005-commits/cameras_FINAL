"""
06_build_workbook.py
--------------------
Assembles the 8-tab Excel workbook that holds the full analysis in
human-readable form. The workbook is the back-pocket reference for the
story — reporters can sort, filter, or re-cut any of these tables without
rerunning the Python pipeline.

Input:
    data/processed/per_camera_results.pkl
    data/processed/placebo_results.pkl

Output:
    outputs/camera_analysis_results.xlsx

Tabs produced (in order):
    1. README         — what's in the workbook and how to read it
    2. Summary        — headline DiD numbers by camera type and outcome
    3. Per_Camera     — one row per camera with all 87 computed columns
    4. Placebo        — real vs placebo DiD side by side
    5. Sensitivity    — DiD at 150m / 200m / 250m
    6. Mechanism      — speeding vs non-speeding crashes near speed cameras
    7. Displacement   — donut-ring comparison (200-500m)
    8. Methodology    — prose notes on design, caveats, limitations
"""

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO_ROOT = Path(__file__).resolve().parent.parent
RESULTS_PATH = REPO_ROOT / "data" / "processed" / "per_camera_results.pkl"
PLACEBO_PATH = REPO_ROOT / "data" / "processed" / "placebo_results.pkl"
OUT_PATH = REPO_ROOT / "outputs" / "camera_analysis_results.xlsx"

TYPES = ["Speed", "Red Light", "Stop Sign", "Truck Restriction"]
OUTCOMES_DISPLAY = [
    ("all", "All crashes"),
    ("injury", "Injury crashes"),
    ("serious", "Serious (fatal+major)"),
    ("fatal", "Fatal only"),
]

# --- Shared visual constants for the workbook.
HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
YELLOW_FILL = PatternFill("solid", start_color="FFF2CC")


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def autosize_columns(ws, min_w=10, max_w=30):
    """Rough column autosizing — openpyxl can't measure true pixel widths."""
    for col in ws.columns:
        cells = [c for c in col if c.column_letter]
        if not cells:
            continue
        length = max(len(str(c.value)) if c.value is not None else 0 for c in cells)
        ws.column_dimensions[cells[0].column_letter].width = min(max(length + 2, min_w), max_w)


def did_row(sub_real, sub_placebo, outcome_key, camera_type, n):
    """Compute one row of summary numbers for a (type, outcome) pair."""
    # --- Real DiD
    pre = sub_real[f"b200_{outcome_key}_pre"].sum()
    post = sub_real[f"b200_{outcome_key}_post"].sum()
    cpre = sub_real[f"citywide_{outcome_key}_pre"].sum()
    cpost = sub_real[f"citywide_{outcome_key}_post"].sum()
    zpct = (post - pre) / pre * 100 if pre else None
    cpct = (cpost - cpre) / cpre * 100 if cpre else None
    did = zpct - cpct if (zpct is not None and cpct is not None) else None
    # --- Placebo DiD
    ppre = sub_placebo[f"placebo_{outcome_key}_pre"].sum()
    ppost = sub_placebo[f"placebo_{outcome_key}_post"].sum()
    pcpre = sub_placebo[f"placebo_city_{outcome_key}_pre"].sum()
    pcpost = sub_placebo[f"placebo_city_{outcome_key}_post"].sum()
    pzpct = (ppost - ppre) / ppre * 100 if ppre else None
    pcpct = (pcpost - pcpre) / pcpre * 100 if pcpre else None
    pdid = pzpct - pcpct if (pzpct is not None and pcpct is not None) else None
    corrected = did - pdid if (did is not None and pdid is not None) else None

    return [camera_type, n,
            round(pre, 1), round(post, 1),
            round(zpct, 1) if zpct is not None else None,
            round(cpct, 1) if cpct is not None else None,
            round(did, 1) if did is not None else None,
            round(pdid, 1) if pdid is not None else None,
            round(corrected, 1) if corrected is not None else None]


def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    for row in [
        ["DC Traffic Camera Impact Analysis"],
        [""],
        ["Analysis window: crash data from 2021-01-02 through 2026-03-02."],
        ["Eligible cameras: 283 live cameras installed Jan 2022 - Mar 2025."],
        [""],
        ["Primary design: Difference-in-Differences (DiD)"],
        ["  Zone % change in crashes near camera  MINUS  Citywide % change"],
        ["  Negative DiD = zone did better than citywide trend"],
        [""],
        ["Defaults:"],
        ["  Buffer: 200m primary (sensitivity at 150m and 250m)"],
        ["  Windows: 365 days before vs 365 days after install (day-of excluded)"],
        ["  Outcomes: All / Injury (HEADLINE) / Serious (fatal+major) / Fatal / Speeding"],
        ["  Camera types: Speed (185), Red Light (56), Stop Sign (32), Truck Restriction (10)"],
        [""],
        ["TABS:"],
        ["  Summary       — Headline DiD numbers by camera type"],
        ["  Per_Camera    — One row per eligible camera, all computed columns"],
        ["  Placebo       — Real vs placebo DiD (install date shifted back 1 yr)"],
        ["  Sensitivity   — DiD at 150m / 200m / 250m buffers"],
        ["  Mechanism     — Speed cameras: speeding-involved vs non-speeding"],
        ["  Displacement  — Donut ring (200-500m) comparison for spillover"],
        ["  Methodology   — Design notes and caveats for the story"],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=16, color="1F3A5F")
    for r in [6, 10, 16]:
        ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 100


def build_summary(wb, res, placebo):
    ws = wb.create_sheet("Summary")
    ws.append(["Headline results by camera type — 200m buffer"])
    ws.append([])
    ws.append(["Camera Type", "N", "Pre total", "Post total", "Zone %",
               "Citywide %", "DiD", "Placebo DiD", "Corrected DiD"])

    for outcome_key, outcome_label in OUTCOMES_DISPLAY:
        ws.append([f"— {outcome_label} —"])
        for t in TYPES:
            sub_r = res[res["type"] == t]
            sub_p = placebo[placebo["type"] == t]
            row = did_row(sub_r, sub_p, outcome_key, t, len(sub_r))
            ws.append(row)
        ws.append([])

    style_header_row(ws, 3)
    ws["A1"].font = Font(bold=True, size=13)

    # --- Color-code the DiD column (index 7) for quick scanning.
    for row_idx in range(4, ws.max_row + 1):
        did_cell = ws.cell(row=row_idx, column=7)
        if did_cell.value is None:
            continue
        try:
            v = float(did_cell.value)
            if v <= -5:
                did_cell.fill = GREEN_FILL
            elif v >= 5:
                did_cell.fill = RED_FILL
            else:
                did_cell.fill = YELLOW_FILL
        except (ValueError, TypeError):
            pass

    autosize_columns(ws, min_w=12, max_w=25)


def build_per_camera(wb, res):
    ws = wb.create_sheet("Per_Camera")
    # --- A curated column order puts the most interesting columns first.
    cols = [
        "camera_code", "location", "type", "install_date", "ward", "lat", "lon",
        "b200_all_pre", "b200_all_post", "b200_all_pct",
        "b200_injury_pre", "b200_injury_post", "b200_injury_pct",
        "b200_serious_pre", "b200_serious_post", "b200_serious_pct",
        "citywide_all_pct", "citywide_injury_pct",
        "did_200m_all", "did_200m_injury", "did_200m_serious",
        "b150_all_pre", "b150_all_post", "b150_injury_pre", "b150_injury_post",
        "b250_all_pre", "b250_all_post", "b250_injury_pre", "b250_injury_post",
        "donut_all_pre", "donut_all_post", "donut_injury_pre", "donut_injury_post",
    ]
    out = res[cols].copy()
    # --- Round all floats to 2 decimals for legibility.
    for c in out.columns:
        if out[c].dtype == float:
            out[c] = out[c].round(2)

    ws.append(list(out.columns))
    for _, row in out.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.values])
    style_header_row(ws)
    ws.freeze_panes = "D2"
    autosize_columns(ws)


def build_placebo(wb, res, placebo):
    ws = wb.create_sheet("Placebo")
    ws.append(["Placebo test — install dates shifted back 1 year"])
    ws.append(["Real DiD should differ from placebo DiD if the effect is real."])
    ws.append([])
    ws.append(["Camera Type", "N", "Pre total", "Post total", "Real Zone%",
               "Real City%", "Real DiD", "Placebo DiD", "Corrected DiD"])
    for outcome_key, outcome_label in OUTCOMES_DISPLAY:
        ws.append([f"— {outcome_label} —"])
        for t in TYPES:
            sub_r = res[res["type"] == t]
            sub_p = placebo[placebo["type"] == t]
            ws.append(did_row(sub_r, sub_p, outcome_key, t, len(sub_r)))
        ws.append([])
    style_header_row(ws, 4)
    ws["A1"].font = Font(bold=True, size=13)
    autosize_columns(ws, min_w=12)


def build_sensitivity(wb, res):
    ws = wb.create_sheet("Sensitivity")
    ws.append(["Sensitivity to buffer radius — injury crashes DiD"])
    ws.append(["Robust findings should not flip direction when the buffer changes."])
    ws.append([])
    ws.append(["Camera Type", "N", "150m DiD", "200m DiD", "250m DiD", "Direction stable?"])

    for t in TYPES:
        sub = res[res["type"] == t]
        row = [t, len(sub)]
        dids = []
        for b in [150, 200, 250]:
            pre = sub[f"b{b}_injury_pre"].sum()
            post = sub[f"b{b}_injury_post"].sum()
            cpre = sub["citywide_injury_pre"].sum()
            cpost = sub["citywide_injury_post"].sum()
            zpct = (post - pre) / pre * 100 if pre else None
            cpct = (cpost - cpre) / cpre * 100 if cpre else None
            d = zpct - cpct if (zpct is not None and cpct is not None) else None
            dids.append(d)
            row.append(round(d, 1) if d is not None else None)
        # --- Stable = all three DiDs have the same sign.
        if all(d is not None for d in dids) and (all(d >= 0 for d in dids) or all(d <= 0 for d in dids)):
            row.append("Yes")
        else:
            row.append("No")
        ws.append(row)
    style_header_row(ws, 4)
    ws["A1"].font = Font(bold=True, size=13)
    autosize_columns(ws, min_w=12)


def build_mechanism(wb, res):
    ws = wb.create_sheet("Mechanism")
    ws.append(["Mechanism check — speed cameras only, 200m buffer"])
    ws.append(["If speed cameras work, speeding-involved crashes should fall MORE than non-speeding."])
    ws.append([])
    ws.append(["Outcome", "Zone pre", "Zone post", "Zone %",
               "City pre", "City post", "City %", "DiD"])

    sub = res[res["type"] == "Speed"]

    def push_row(label, zp, zpo, cp, cpo):
        zpct = (zpo - zp) / zp * 100 if zp else None
        cpct = (cpo - cp) / cp * 100 if cp else None
        did = zpct - cpct if (zpct is not None and cpct is not None) else None
        ws.append([label, zp, zpo,
                   round(zpct, 1) if zpct is not None else None,
                   cp, cpo,
                   round(cpct, 1) if cpct is not None else None,
                   round(did, 1) if did is not None else None])

    push_row("Speeding-involved crashes",
             sub["b200_speeding_pre"].sum(), sub["b200_speeding_post"].sum(),
             sub["citywide_speeding_pre"].sum(), sub["citywide_speeding_post"].sum())
    push_row("All crashes",
             sub["b200_all_pre"].sum(), sub["b200_all_post"].sum(),
             sub["citywide_all_pre"].sum(), sub["citywide_all_post"].sum())
    # --- Non-speeding = all minus speeding, computed at the aggregate level.
    push_row("Non-speeding crashes",
             sub["b200_all_pre"].sum() - sub["b200_speeding_pre"].sum(),
             sub["b200_all_post"].sum() - sub["b200_speeding_post"].sum(),
             sub["citywide_all_pre"].sum() - sub["citywide_speeding_pre"].sum(),
             sub["citywide_all_post"].sum() - sub["citywide_speeding_post"].sum())

    style_header_row(ws, 4)
    ws["A1"].font = Font(bold=True, size=13)
    autosize_columns(ws, min_w=14)


def build_displacement(wb, res):
    ws = wb.create_sheet("Displacement")
    ws.append(["Displacement check — 200-500m donut ring"])
    ws.append(["If cameras only push bad driving elsewhere, the donut should RISE after install."])
    ws.append([])
    ws.append(["Camera Type", "N", "Zone 200m %", "Donut 200-500m %", "Displacement evidence?"])

    for t in TYPES:
        sub = res[res["type"] == t]
        zpre, zpost = sub["b200_injury_pre"].sum(), sub["b200_injury_post"].sum()
        dpre, dpost = sub["donut_injury_pre"].sum(), sub["donut_injury_post"].sum()
        zpct = (zpost - zpre) / zpre * 100 if zpre else None
        dpct = (dpost - dpre) / dpre * 100 if dpre else None
        # --- Evidence = zone fell AND donut rose (the displacement pattern).
        if zpct is not None and dpct is not None and zpct < 0 and dpct > 0:
            evidence = "Yes (zone fell, donut rose)"
        else:
            evidence = "No"
        ws.append([t, len(sub),
                   round(zpct, 1) if zpct is not None else None,
                   round(dpct, 1) if dpct is not None else None,
                   evidence])
    style_header_row(ws, 4)
    ws["A1"].font = Font(bold=True, size=13)
    autosize_columns(ws, min_w=14)


def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 110
    sections = [
        ("DC Traffic Camera Impact Analysis — Methodology", True, 15),
        ("", False, 11),
        ("DATA", True, 12),
        ("Crash data: DDOT Crashes in DC dataset (75,400 geocoded crashes, 2021-2026).", False, 11),
        ("Camera data: DDOT Automated Traffic Enforcement dataset (283 eligible cameras).", False, 11),
        ("", False, 11),
        ("DESIGN", True, 12),
        ("Primary metric: Difference-in-Differences (DiD) = zone % change − citywide % change.", False, 11),
        ("Buffer: 200m Euclidean radius, proxy for 'two DC blocks'. Sensitivity at 150m / 250m.", False, 11),
        ("Windows: 365 days before install vs 365 days after (install day excluded both sides).", False, 11),
        ("", False, 11),
        ("VALIDITY CHECKS", True, 12),
        ("Placebo: install dates shifted back 1 year. Real effects should produce near-zero placebo.", False, 11),
        ("Sensitivity: DiD recomputed at 150m / 200m / 250m. Robust findings agree in direction.", False, 11),
        ("Mechanism: speed cameras — speeding-involved vs non-speeding crashes.", False, 11),
        ("Displacement: 200-500m donut ring. If cameras only displace crashes, donut rises.", False, 11),
        ("", False, 11),
        ("LIMITATIONS", True, 12),
        ("1. Non-random camera placement creates regression-to-the-mean risk.", False, 11),
        ("2. Euclidean buffer is a proxy for street-network distance.", False, 11),
        ("3. Camera installs may coincide with other streetscape changes (signs, paint, bollards).", False, 11),
        ("4. START_DATE is assumed to mean active enforcement; warning periods would dilute the effect.", False, 11),
        ("5. Fatal/serious per-camera-type counts are too sparse for confident claims.", False, 11),
        ("6. Overlapping buffers between nearby cameras are not de-duplicated at the aggregate level.", False, 11),
    ]
    for text, bold, size in sections:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=bold, size=size, name="Arial")
        c.alignment = Alignment(wrap_text=True)


def main():
    print("Loading per-camera + placebo results ...")
    res = pd.read_pickle(RESULTS_PATH)
    placebo = pd.read_pickle(PLACEBO_PATH)

    # --- Attach the camera type to the placebo rows for groupbys later.
    placebo = placebo.merge(res[["camera_code", "type"]].rename(columns={"type": "_type"}),
                            on="camera_code", how="left")
    if "type" not in placebo.columns:
        placebo = placebo.rename(columns={"_type": "type"})

    wb = Workbook()
    build_readme(wb)
    build_summary(wb, res, placebo)
    build_per_camera(wb, res)
    build_placebo(wb, res, placebo)
    build_sensitivity(wb, res)
    build_mechanism(wb, res)
    build_displacement(wb, res)
    build_methodology(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nWrote {OUT_PATH.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
