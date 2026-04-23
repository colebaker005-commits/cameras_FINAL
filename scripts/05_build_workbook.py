# updated Script 6, checked by Cole
"""

06_build_workbook.py
--------------------
Assembles the final Excel workbook from the analysis results. This is the
human-readable version of everything computed in scripts 03 and 04.

Tabs produced:
    1. README        — what's in the workbook and how to read it
    2. Summary       — headline DiD numbers by camera type
    3. Per_Camera    — one row per camera with all computed columns
    4. Placebo       — real DiD vs placebo DiD side by side
    5. Methodology   — notes on design, caveats, limitations

Input:
    data/processed/per_camera_results.pkl
    data/processed/placebo_results.pkl

Output:
    outputs/camera_analysis_results.xlsx
"""

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPO_ROOT    = Path(__file__).resolve().parent.parent
RESULTS_PATH = REPO_ROOT / "data" / "processed" / "per_camera_results.pkl"
PLACEBO_PATH = REPO_ROOT / "data" / "processed" / "placebo_results.pkl"
OUT_PATH     = REPO_ROOT / "outputs" / "camera_analysis_results.xlsx"

TYPES    = ["Speed", "Red Light", "Stop Sign", "Truck Restriction"]
OUTCOMES = [
    ("all",     "All crashes"),
    ("injury",  "Injury crashes"),
    ("serious", "Serious (fatal + major)"),
    ("fatal",   "Fatal only"),
]

HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
RED_FILL    = PatternFill("solid", start_color="FFC7CE")
YELLOW_FILL = PatternFill("solid", start_color="FFF2CC")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def autosize(ws, min_w=10, max_w=30):
    for col in ws.columns:
        cells = [c for c in col if c.column_letter]
        if not cells:
            continue
        length = max(len(str(c.value)) if c.value is not None else 0 for c in cells)
        ws.column_dimensions[cells[0].column_letter].width = min(max(length + 2, min_w), max_w)


def compute_did(sub, nearby_pre, nearby_post, city_pre, city_post):
    pre   = sub[nearby_pre].sum()
    post  = sub[nearby_post].sum()
    cpre  = sub[city_pre].sum()
    cpost = sub[city_post].sum()
    zone  = (post - pre)  / pre  * 100 if pre  else None
    city  = (cpost - cpre) / cpre * 100 if cpre else None
    if zone is None or city is None:
        return None
    return round(zone - city, 1)





def build_summary(wb, res, placebo):
    ws = wb.active
    ws.append(["Headline results by camera type — 200m buffer"])
    ws.append([])
    ws.append(["Camera Type", "N", "Outcome",
               "Nearby pre", "Nearby post", "Nearby %",
               "Citywide %", "DiD",
               "Placebo DiD", "Corrected DiD (Real - Placebo)"])

    for outcome_key, outcome_label in OUTCOMES:
        for t in TYPES:
            sub_r = res[res["type"] == t]
            sub_p = placebo[placebo["type"] == t]

            # Real DiD
            real_did = compute_did(sub_r,
                f"nearby_{outcome_key}_pre", f"nearby_{outcome_key}_post",
                f"citywide_{outcome_key}_pre", f"citywide_{outcome_key}_post")

            # Nearby and citywide raw numbers
            npre  = sub_r[f"nearby_{outcome_key}_pre"].sum()
            npost = sub_r[f"nearby_{outcome_key}_post"].sum()
            cpre  = sub_r[f"citywide_{outcome_key}_pre"].sum()
            cpost = sub_r[f"citywide_{outcome_key}_post"].sum()
            npct  = round((npost - npre) / npre * 100, 1) if npre else None
            cpct  = round((cpost - cpre) / cpre * 100, 1) if cpre else None

            # Placebo DiD
            placebo_did = compute_did(sub_p,
                f"placebo_nearby_{outcome_key}_pre", f"placebo_nearby_{outcome_key}_post",
                f"placebo_city_{outcome_key}_pre",   f"placebo_city_{outcome_key}_post")

            corrected = round(real_did - placebo_did, 1) if (real_did is not None and placebo_did is not None) else None

            ws.append([t, len(sub_r), outcome_label,
                       npre, npost, npct, cpct,
                       real_did, placebo_did, corrected])

        ws.append([])

    style_header(ws, 3)
    ws["A1"].font = Font(bold=True, size=13)

    # Color code the DiD column (col 8)
    for row_idx in range(4, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=8)
        if cell.value is None:
            continue
        try:
            v = float(cell.value)
            if v <= -5:
                cell.fill = GREEN_FILL
            elif v >= 5:
                cell.fill = RED_FILL
            else:
                cell.fill = YELLOW_FILL
        except (ValueError, TypeError):
            pass

    autosize(ws, min_w=12, max_w=25)


def build_per_camera(wb, res):
    ws = wb.create_sheet("Per_Camera")

    cols = [
        "camera_code", "location", "type", "install_date", "ward", "lat", "lon",
        "nearby_all_pre",     "nearby_all_post",     "nearby_all_pct",
        "nearby_injury_pre",  "nearby_injury_post",  "nearby_injury_pct",
        "nearby_serious_pre", "nearby_serious_post", "nearby_serious_pct",
        "nearby_fatal_pre",   "nearby_fatal_post",   "nearby_fatal_pct",
        "citywide_all_pct",   "citywide_injury_pct",
        "did_all", "did_injury", "did_serious", "did_fatal", "did_speeding",
    ]
    out = res[cols].copy()
    for c in out.columns:
        if out[c].dtype == float:
            out[c] = out[c].round(2)

    ws.append(list(out.columns))
    for _, row in out.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.values])

    style_header(ws)
    ws.freeze_panes = "D2"
    autosize(ws)


def build_placebo(wb, res, placebo):
    ws = wb.create_sheet("Placebo")
    ws.append(["Placebo test — install dates shifted back 1 year"])
    ws.append(["If the real DiD is genuinely caused by the camera, the placebo DiD should be close to zero."])
    ws.append([])
    ws.append(["Camera Type", "N", "Outcome",
               "Real DiD", "Placebo DiD", "Corrected DiD (Real - Placebo)"])

    for outcome_key, outcome_label in OUTCOMES:
        for t in TYPES:
            sub_r = res[res["type"] == t]
            sub_p = placebo[placebo["type"] == t]

            real_did = compute_did(sub_r,
                f"nearby_{outcome_key}_pre", f"nearby_{outcome_key}_post",
                f"citywide_{outcome_key}_pre", f"citywide_{outcome_key}_post")

            placebo_did = compute_did(sub_p,
                f"placebo_nearby_{outcome_key}_pre", f"placebo_nearby_{outcome_key}_post",
                f"placebo_city_{outcome_key}_pre",   f"placebo_city_{outcome_key}_post")

            corrected = round(real_did - placebo_did, 1) if (real_did is not None and placebo_did is not None) else None

            ws.append([t, len(sub_r), outcome_label,
                       real_did, placebo_did, corrected])
        ws.append([])

    style_header(ws, 4)
    ws["A1"].font = Font(bold=True, size=13)
    autosize(ws, min_w=12)





def main():
    print("Loading results ...")
    res     = pd.read_pickle(RESULTS_PATH)
    placebo = pd.read_pickle(PLACEBO_PATH)
    print(f"  {len(res)} cameras in results, {len(placebo)} in placebo")

    wb = Workbook()

    # --- Rename the default sheet and use it for Summary
    wb.active.title = "Summary"
    build_summary(wb, res, placebo)
    build_per_camera(wb, res)
    build_placebo(wb, res, placebo)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nWrote {OUT_PATH.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()